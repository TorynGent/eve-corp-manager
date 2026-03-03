#!/usr/bin/env python3
"""
Import Alt → Main mappings from Corp Management.xlsx into the EVE Dashboard database.

Usage:
    python import_alts.py

Requirements:
    pip install openpyxl
"""

import sys
import os
import sqlite3

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Run:   pip install openpyxl")
    sys.exit(1)

# ── Adjust these paths if needed ──────────────────────────────────────────────
XLSX_PATH  = r"G:\Thumbnails\EVE\Corp Management.xlsx"
DB_PATH    = r"G:\Thumbnails\EVE\eve-app\data\corp.db"
SHEET_NAME = "Settings"   # Sheet that contains the alt→main data
# ─────────────────────────────────────────────────────────────────────────────


def find_columns(headers):
    """Return (alt_col_index, main_col_index) based on header names."""
    alt_col = main_col = None
    for i, h in enumerate(headers):
        h_lower = str(h or '').strip().lower()
        if alt_col is None and ('alt' in h_lower or 'character' in h_lower):
            alt_col = i
        if main_col is None and 'main' in h_lower:
            main_col = i
    return alt_col, main_col


def hash_name(name):
    """Deterministic integer ID for a name string (negative, for unresolved chars)."""
    h = 0
    for c in name:
        h = (31 * h + ord(c)) & 0xFFFFFFFF
    return -(h % 999_999_998 + 1)


def main():
    # Validate paths
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: Excel file not found:\n  {XLSX_PATH}")
        sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"ERROR: Database not found:\n  {DB_PATH}")
        print("Start the app at least once so the database is created, then re-run this script.")
        sys.exit(1)

    print(f"Opening: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # Pick sheet
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        print(f"Using sheet: '{SHEET_NAME}'")
    else:
        print(f"Sheet '{SHEET_NAME}' not found.")
        print(f"Available sheets: {wb.sheetnames}")
        choice = input("Enter sheet name to use (or press Enter to cancel): ").strip()
        if not choice or choice not in wb.sheetnames:
            print("Cancelled.")
            sys.exit(0)
        ws = wb[choice]

    # Read headers
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        print("ERROR: Sheet is empty.")
        sys.exit(1)

    headers = [str(c or '').strip() for c in header_row]
    print(f"Columns: {headers}")

    alt_col, main_col = find_columns(headers)

    if alt_col is None or main_col is None:
        print(f"\nCould not auto-detect columns from headers: {headers}")
        print("Enter column numbers (0 = first column):")
        try:
            alt_col  = int(input("  Alt character column index: "))
            main_col = int(input("  Main character column index: "))
        except ValueError:
            print("Invalid input.")
            sys.exit(1)
    else:
        print(f"Detected: alt=col {alt_col} ('{headers[alt_col]}'), "
              f"main=col {main_col} ('{headers[main_col]}')")

    # Connect to DB and load name cache for ID resolution
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    cur.execute("SELECT LOWER(name), id FROM name_cache WHERE type = 'character'")
    name_to_id = dict(cur.fetchall())
    print(f"Loaded {len(name_to_id)} cached character names from database.\n")

    imported = 0
    skipped  = 0

    for row_num, row in enumerate(rows_iter, start=2):
        if not row:
            continue
        max_col = max(alt_col, main_col)
        if len(row) <= max_col:
            skipped += 1
            continue

        alt_name  = str(row[alt_col]  or '').strip()
        main_name = str(row[main_col] or '').strip()

        if not alt_name or not main_name or alt_name.lower() == 'none':
            skipped += 1
            continue

        # Resolve character ID from name cache, else use deterministic negative ID
        char_id = name_to_id.get(alt_name.lower(), hash_name(alt_name))

        cur.execute("""
            INSERT INTO alt_mappings (character_id, character_name, main_name)
            VALUES (?, ?, ?)
            ON CONFLICT(character_id) DO UPDATE SET
                character_name = excluded.character_name,
                main_name      = excluded.main_name
        """, (char_id, alt_name, main_name))

        print(f"  [{row_num}] {alt_name} → {main_name}  (charId: {char_id})")
        imported += 1

    conn.commit()
    conn.close()

    print(f"\n✓ Done!  Imported: {imported}  |  Skipped: {skipped}")
    print("Go to Settings → Sync Now in the dashboard to refresh aggregated data.")


if __name__ == '__main__':
    main()
